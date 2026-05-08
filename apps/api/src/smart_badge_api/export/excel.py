"""将分析任务导出为 Excel 工作簿。"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from smart_badge_api.api.analysis_normalization import normalize_analysis_result
from smart_badge_api.db.models import AnalysisTask

# ── 样式常量 ──────────────────────────────────────────────────
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_WRAP = Alignment(vertical="top", wrap_text=True)


def _style_header(ws, row: int = 1):
    """给表头行添加样式。"""
    for cell in ws[row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _auto_width(ws, min_width: int = 12, max_width: int = 50):
    """自动调整列宽。"""
    for col in ws.columns:
        best = min_width
        for cell in col:
            if cell.value:
                best = max(best, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col[0].column_letter].width = best + 2


def _fmt_dt(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")


# ── 单任务导出 ────────────────────────────────────────────────

def export_task_excel(task: AnalysisTask) -> bytes:
    """将单个任务导出为 Excel 工作簿 (bytes)。"""
    wb = Workbook()

    # ── Sheet 1: 概览 ──
    ws_summary = wb.active
    ws_summary.title = "任务概览"
    ws_summary.append(["字段", "值"])
    _style_header(ws_summary)
    rows = [
        ("文件名", task.file_name),
        ("状态", task.status),
        ("进度", f"{task.progress}%"),
        ("综合评分", task.overall_score if task.overall_score is not None else "N/A"),
        ("音频时长(毫秒)", task.duration_ms or "N/A"),
        ("语音片段数", task.segment_count or "N/A"),
        ("错误信息", task.error_message or ""),
        ("创建时间", _fmt_dt(task.created_at)),
        ("完成时间", _fmt_dt(task.completed_at)),
    ]
    for r in rows:
        ws_summary.append(r)
    _auto_width(ws_summary)

    result: dict | None = normalize_analysis_result(task.result) if isinstance(task.result, dict) else task.result
    if not result:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Sheet 2: 接诊评价 ──
    eval_data = result.get("consultation_evaluation", {})
    ws_eval = wb.create_sheet("接诊评价")
    ws_eval.append(["维度名称", "评分", "评语"])
    _style_header(ws_eval)
    for dim in eval_data.get("dimensions", []):
        ws_eval.append([dim.get("name", ""), dim.get("score", ""), dim.get("comment", "")])
    if eval_data.get("overall_score") is not None:
        ws_eval.append([])
        ws_eval.append(["综合评分", eval_data["overall_score"], ""])
    _auto_width(ws_eval)

    # ── Sheet 3: 客户诉求 ──
    demands = result.get("customer_demands", {})
    ws_demands = wb.create_sheet("客户诉求")
    ws_demands.append(["部位", "表层需求", "深层需求", "需求挖掘过程"])
    _style_header(ws_demands)
    for fa in demands.get("focus_areas", []):
        ws_demands.append([
            fa.get("area", ""),
            fa.get("surface_need", ""),
            fa.get("deep_need", ""),
            fa.get("discovery_process", ""),
        ])

    # 期望轨迹
    exp = demands.get("expectation", {})
    if exp:
        ws_demands.append([])
        ws_demands.append(["期望效果与心态变化"])
        ws_demands.append(["对话类型", exp.get("dialogue_type", "")])
        ws_demands.append(["入口状态", exp.get("entry_state", "")])
        ws_demands.append(["出口状态", exp.get("exit_state", "")])
        turning = exp.get("turning_points", [])
        if turning:
            ws_demands.append(["关键转折点", "; ".join(turning)])
        ws_demands.append(["效果标准", exp.get("specific_standards", "")])

    # 产品倾向
    pref = demands.get("product_preference", {})
    if pref:
        ws_demands.append([])
        ws_demands.append(["产品倾向分析"])
        ws_demands.append(["倾向产品", ", ".join(pref.get("preferred_products", []))])
        ws_demands.append(["信息来源", ", ".join(pref.get("information_sources", []))])
        ws_demands.append(["比较因素", ", ".join(pref.get("comparison_factors", []))])
        ws_demands.append(["咨询师引导", pref.get("consultant_influence", "")])
    _auto_width(ws_demands)

    # ── Sheet 4: 顾客顾虑 ──
    concerns = result.get("customer_concerns", {})
    ws_concerns = wb.create_sheet("顾客顾虑")
    if concerns.get("summary"):
        ws_concerns.append(["概述"])
        ws_concerns.append([concerns["summary"]])
        ws_concerns.append([])
    ws_concerns.append(["类别", "内容", "证据原文"])
    _style_header(ws_concerns, ws_concerns.max_row)
    for item in concerns.get("items", []):
        ws_concerns.append([item.get("type", ""), item.get("content", ""), item.get("evidence", "")])
    _auto_width(ws_concerns)

    # ── Sheet 5: 客户画像 ──
    profile = result.get("customer_profile", {})
    ws_profile = wb.create_sheet("客户画像")
    ws_profile.append(["标签分类", "标签值"])
    _style_header(ws_profile)
    for tag in profile.get("tags", []):
        ws_profile.append([tag.get("category", ""), tag.get("value", "")])
    _auto_width(ws_profile)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 批量导出 ──────────────────────────────────────────────────

def export_tasks_batch_excel(tasks: list[AnalysisTask]) -> bytes:
    """将多个任务汇总导出为 Excel（每行一个任务）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "任务汇总"
    ws.append([
        "任务ID", "文件名", "状态", "综合评分",
        "对话类型", "重点部位数", "顾虑条数", "标签数",
        "音频时长(ms)", "语音片段数", "创建时间", "完成时间",
    ])
    _style_header(ws)

    for t in tasks:
        result = normalize_analysis_result(t.result) if isinstance(t.result, dict) else (t.result or {})
        demands = result.get("customer_demands", {})
        concerns = result.get("customer_concerns", {})
        profile = result.get("customer_profile", {})
        exp = demands.get("expectation", {})

        ws.append([
            t.id,
            t.file_name,
            t.status,
            t.overall_score if t.overall_score is not None else "",
            exp.get("dialogue_type", ""),
            len(demands.get("focus_areas", [])),
            len(concerns.get("items", [])),
            len(profile.get("tags", [])),
            t.duration_ms or "",
            t.segment_count or "",
            _fmt_dt(t.created_at),
            _fmt_dt(t.completed_at),
        ])

    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
